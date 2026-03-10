"""
主窗口界面 V2 - 参考 VoiceInk 设计
左侧边栏 + 右侧内容区布局
"""
import os
import time
import csv
from datetime import datetime
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QLineEdit, QPushButton, QTextEdit, 
    QStackedWidget, QProgressBar, QFileDialog, QMessageBox, 
    QSpinBox, QFormLayout, QFrame, QCheckBox, QComboBox,
    QScrollArea, QSizePolicy, QGridLayout, QDialog, QListWidget
)
from PyQt6.QtCore import QSettings, Qt, QTimer, QRectF
from PyQt6.QtGui import QFont, QPainter, QColor, QBrush, QPen
from ..core import EmailSender, ContactFetcher, ConfigManager
from ..utils import read_contacts
from .styles_premium import PURPLE_THEME
from .custom_widgets import PremiumSpinBox
from .contact_manager import ContactManagerWidget
from .quick_send import QuickSendDialog, ContactSelectDialog


class SidebarButton(QPushButton):
    """侧边栏按钮"""
    def __init__(self, text: str, icon: str = "", parent=None):
        super().__init__(parent)
        self.setText(f"{icon}  {text}" if icon else text)
        self.setCheckable(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(44)


class BarChart(QWidget):
    """简单的条形图组件"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = {}  # {label: value}
        self.setMinimumHeight(180)
        self.colors = [
            QColor("#6C5CE7"),  # Theme Primary
            QColor("#A371F7"),  # Purple Light
            QColor("#10B981"),  # Emerald
            QColor("#F59E0B"),  # Amber
            QColor("#EF4444"),  # Red
            QColor("#8B5CF6"),  # Violet
            QColor("#EC4899"),  # Pink
            QColor("#6366F1"),  # Indigo
        ]
    
    def set_data(self, data: dict):
        """设置数据 {label: value}"""
        self.data = data
        self.update()
    
    def paintEvent(self, event):
        if not self.data:
            return
        
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        width = self.width()
        height = self.height()
        padding = 20
        bar_spacing = 8
        
        # 计算最大值
        max_val = max(self.data.values()) if self.data else 1
        
        # 取前8个数据
        items = list(self.data.items())[:8]
        if not items:
            return
        
        bar_width = min(50, (width - padding * 2 - bar_spacing * (len(items) - 1)) / len(items))
        chart_height = height - 60  # 留空间给标签
        
        # 绘制条形
        x = padding + (width - padding * 2 - bar_width * len(items) - bar_spacing * (len(items) - 1)) / 2
        
        for i, (label, value) in enumerate(items):
            # 条形高度
            bar_height = (value / max_val) * chart_height if max_val > 0 else 0
            bar_height = max(bar_height, 4)  # 最小高度
            
            # 颜色
            color = self.colors[i % len(self.colors)]
            
            # 绘制条形（圆角矩形）
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.PenStyle.NoPen)
            
            rect_x = int(x)
            rect_y = int(height - 50 - bar_height)
            rect_w = int(bar_width)
            rect_h = int(bar_height)
            
            painter.drawRoundedRect(rect_x, rect_y, rect_w, rect_h, 4, 4)
            
            # 绘制数值
            painter.setPen(QPen(QColor("#FFFFFF")))
            font = painter.font()
            font.setPointSize(10)
            font.setBold(True)
            painter.setFont(font)
            painter.drawText(rect_x, rect_y - 5, rect_w, 20, 
                           Qt.AlignmentFlag.AlignCenter, str(value))
            
            # 绘制标签
            painter.setFont(font)
            
            # 截断过长的标签
            display_label = label[:8] + ".." if len(label) > 10 else label
            painter.drawText(rect_x - 5, int(height - 35), rect_w + 10, 30, 
                           Qt.AlignmentFlag.AlignCenter, display_label)
            
            x += bar_width + bar_spacing


class PieChart(QWidget):
    """饼状图组件 (Donut style)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = {}  # {label: value}
        self.setMinimumHeight(220) # Slightly taller for legend
        self.colors = [
            QColor("#6C5CE7"),  # Theme Primary
            QColor("#A371F7"),  # Purple Light
            QColor("#10B981"),  # Emerald
            QColor("#F59E0B"),  # Amber
            QColor("#EF4444"),  # Red
            QColor("#8B5CF6"),  # Violet
            QColor("#EC4899"),  # Pink
            QColor("#6366F1"),  # Indigo
        ]
    
    def set_data(self, data: dict):
        self.data = data
        self.update()
    
    def paintEvent(self, event):
        if not self.data:
            return
        
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        width = self.width()
        height = self.height()
        
        # Calculate total
        total = sum(self.data.values())
        if total == 0:
            return
            
        # Layout: Chart Left, Legend Right
        chart_size = min(width * 0.5, height - 20)
        chart_rect = QRectF(20, (height - chart_size) / 2, chart_size, chart_size)
        
        # Draw Pie/Donut
        start_angle = 90 * 16 # Start from top
        
        items = list(self.data.items())[:8] # Top 8
        other_val = total - sum(v for k, v in items)
        if other_val > 0:
            items.append(("其他", other_val))
            
        used_colors = []
        
        for i, (label, value) in enumerate(items):
            span_angle = -(value / total) * 360 * 16
            color = self.colors[i % len(self.colors)]
            used_colors.append(color)
            
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawPie(chart_rect, int(start_angle), int(span_angle))
            
            start_angle += span_angle
            
        # Draw Center Hole (Donut)
        hole_size = chart_size * 0.6
        hole_rect = QRectF(
            chart_rect.center().x() - hole_size/2,
            chart_rect.center().y() - hole_size/2,
            hole_size, hole_size
        )
        painter.setBrush(QBrush(QColor("#1E202E"))) # Match card background
        painter.drawEllipse(hole_rect)
        
        # Draw Legend
        legend_x = 20 + chart_size + 40
        legend_y = (height - len(items) * 24) / 2
        
        painter.setFont(QFont("Arial", 12))
        
        for i, (label, value) in enumerate(items):
            # Color dot
            dot_rect = QRectF(legend_x, legend_y + i*24 + 4, 12, 12)
            painter.setBrush(QBrush(used_colors[i]))
            painter.drawEllipse(dot_rect)
            
            # Text
            painter.setPen(QPen(QColor("#E0E0E0")))
            percent = (value / total) * 100
            text = f"{label} ({percent:.1f}%)"
            painter.drawText(int(legend_x + 20), int(legend_y + i*24 + 14), text)


class CodeLogWidget(QTextEdit):
    """科技感代码日志组件"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self.setMinimumHeight(150)
        self.setMaximumHeight(200)
        
        # 设置等宽字体和样式
        self.setStyleSheet("""
            QTextEdit {
                background-color: #0D1117;
                color: #58A6FF;
                border: 1px solid #30363D;
                border-radius: 8px;
                font-family: 'SF Mono', 'Monaco', 'Menlo', 'Consolas', monospace;
                font-size: 12px;
                padding: 12px;
                line-height: 1.5;
            }
        """)
        
        self.log_index = 0
    
    def add_log(self, message: str, log_type: str = "info"):
        """添加日志，带颜色标记"""
        self.log_index += 1
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        color_map = {
            "info": "#58A6FF",     # 蓝色
            "success": "#3FB950",   # 绿色
            "warning": "#D29922",   # 黄色
            "error": "#F85149",     # 红色
            "data": "#A371F7",      # 紫色
            "progress": "#79C0FF",  # 浅蓝
        }
        color = color_map.get(log_type, "#8B949E")
        
        # 添加带颜色的HTML格式日志
        html = f'<span style="color: #6E7681;">[{timestamp}]</span> '
        html += f'<span style="color: #8B949E;">#{self.log_index:03d}</span> '
        html += f'<span style="color: {color};">{message}</span>'
        
        self.append(html)
        
        # 滚动到底部
        scrollbar = self.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def clear_logs(self):
        """清空日志"""
        self.clear()
        self.log_index = 0


class StatsCard(QFrame):
    """统计卡片 - 左右布局"""
    def __init__(self, icon: str, title: str, value: str, subtitle: str = "", parent=None):
        super().__init__(parent)
        self.setObjectName("statsCard")
        self.setMinimumHeight(100)
        
        # 主布局：左右分布
        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(16)
        
        # 左侧：图标 + 文字信息
        left_layout = QVBoxLayout()
        left_layout.setSpacing(4)
        
        # 图标 + 标题行
        header = QHBoxLayout()
        header.setSpacing(8)
        
        icon_label = QLabel(icon)
        icon_label.setObjectName("cardIcon")
        icon_label.setFixedSize(28, 28)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setObjectName("cardTitle")
        header.addWidget(title_label)
        header.addStretch()
        
        left_layout.addLayout(header)
        
        # 副标题
        if subtitle:
            sub_label = QLabel(subtitle)
            sub_label.setObjectName("cardSubtitle")
            left_layout.addWidget(sub_label)
        
        left_layout.addStretch()
        layout.addLayout(left_layout, 1)
        
        # 右侧：大数字
        self.value_label = QLabel(value)
        self.value_label.setObjectName("cardValue")
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(self.value_label)
    
    def update_value(self, value: str):
        self.value_label.setText(value)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Eason - Email Assistant")
        self.setMinimumSize(1000, 700)
        self.resize(1100, 750)
        
        # 设置macOS深色标题栏
        import platform
        if platform.system() == 'Darwin':  # macOS
            try:
                # 使用pyobjc设置深色主题
                from Foundation import NSBundle
                from AppKit import NSApplication, NSAppearance, NSAppearanceNameDarkAqua
                app = NSApplication.sharedApplication()
                app.setAppearance_(NSAppearance.appearanceNamed_(NSAppearanceNameDarkAqua))
            except ImportError:
                # 如果没有pyobjc，尝试备用方案
                pass
            except Exception:
                pass
        
        # 居中显示
        from PyQt6.QtGui import QScreen
        screen = QScreen.availableGeometry(self.screen())
        x = (screen.width() - 1100) // 2
        y = (screen.height() - 750) // 2
        self.move(x, y)
        
        # 数据
        self.config_manager = ConfigManager()
        self.contacts_data = []
        self.total_sent = 0
        self.success_count = 0
        self.today_sent = 0
        
        # 主题
        # self.current_theme = self.config_manager.load_theme()
        
        # 线程
        self.fetch_thread = None
        self.send_thread = None

        self.init_ui()
        self.apply_theme()
        self.load_config()
        self.load_stats()

    def init_ui(self):
        # 主容器
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ========== 左侧边栏 ==========
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(200)
        
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 20, 12, 20)
        sidebar_layout.setSpacing(4)
        
        # Logo
        logo_container = QWidget()
        logo_layout = QHBoxLayout(logo_container)
        logo_layout.setContentsMargins(12, 0, 0, 20)
        
        logo_text = QLabel("Eason")
        logo_text.setObjectName("logoText")
        
        logo_layout.addWidget(logo_text)
        logo_layout.addStretch()
        sidebar_layout.addWidget(logo_container)
        
        # 导航按钮
        self.nav_buttons = []
        
        self.btn_dashboard = SidebarButton("控制台", "📊")
        self.btn_dashboard.setChecked(True)
        self.btn_dashboard.clicked.connect(lambda: self.switch_page(0))
        sidebar_layout.addWidget(self.btn_dashboard)
        self.nav_buttons.append(self.btn_dashboard)
        
        self.btn_contacts = SidebarButton("联系人管理", "📇")
        self.btn_contacts.clicked.connect(lambda: self.switch_page(1))
        sidebar_layout.addWidget(self.btn_contacts)
        self.nav_buttons.append(self.btn_contacts)
        
        self.btn_collect = SidebarButton("采集联系人", "📥")
        self.btn_collect.clicked.connect(lambda: self.switch_page(2))
        sidebar_layout.addWidget(self.btn_collect)
        self.nav_buttons.append(self.btn_collect)
        
        self.btn_send = SidebarButton("批量发送", "📤")
        self.btn_send.clicked.connect(lambda: self.switch_page(3))
        sidebar_layout.addWidget(self.btn_send)
        self.nav_buttons.append(self.btn_send)
        
        self.btn_settings = SidebarButton("账号设置", "⚙️")
        self.btn_settings.clicked.connect(lambda: self.switch_page(4))
        sidebar_layout.addWidget(self.btn_settings)
        self.nav_buttons.append(self.btn_settings)
        
        sidebar_layout.addStretch()
        
        # 主题切换 (已移除单主题)
        # self.theme_button = SidebarButton("", "🌙")
        # self.theme_button.setCheckable(False)
        # self.theme_button.setObjectName("themeButton")
        # self.update_theme_button_text()
        # self.theme_button.clicked.connect(self.toggle_theme)
        # sidebar_layout.addWidget(self.theme_button)
        
        main_layout.addWidget(sidebar)
        
        # ========== 右侧内容区 ==========
        content_area = QFrame()
        content_area.setObjectName("contentArea")
        
        content_layout = QVBoxLayout(content_area)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        
        # 页面堆栈
        self.pages = QStackedWidget()
        
        # Page 0: 控制台
        self.pages.addWidget(self.create_dashboard_page())
        
        # Page 1: 联系人管理
        self.contact_manager = ContactManagerWidget()
        self.contact_manager.contacts_selected_for_send.connect(self.on_contacts_selected_for_send)
        self.contact_manager.quick_send_requested.connect(self.on_quick_send_requested)
        self.pages.addWidget(self.contact_manager)
        
        # Page 2: 采集联系人
        self.pages.addWidget(self.create_collect_page())
        
        # Page 3: 批量发送
        self.pages.addWidget(self.create_send_page())
        
        # Page 4: 账号设置
        self.pages.addWidget(self.create_settings_page())
        
        content_layout.addWidget(self.pages)
        main_layout.addWidget(content_area, 1)

    def create_dashboard_page(self):
        """创建控制台页面"""
        page = QWidget()
        page.setObjectName("contentArea")
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)
        
        # 页面标题
        title = QLabel("控制台")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        
        # 横幅卡片
        banner = QFrame()
        banner.setObjectName("bannerCard")
        banner_layout = QVBoxLayout(banner)
        banner_layout.setContentsMargins(24, 24, 24, 24)
        banner_layout.setSpacing(8)
        
        banner_title = QLabel("欢迎使用 Eason 邮件助手")
        banner_title.setObjectName("bannerTitle")
        banner_sub = QLabel("高效、稳定的邮件批量发送解决方案")
        banner_sub.setObjectName("bannerSubtitle")
        
        banner_layout.addWidget(banner_title)
        banner_layout.addWidget(banner_sub)
        layout.addWidget(banner)
        
        # 统计卡片网格
        stats_grid = QGridLayout()
        stats_grid.setSpacing(16)
        
        self.card_total = StatsCard("📨", "总发送", "0", "累计发送邮件数")
        self.card_success = StatsCard("✅", "成功率", "0%", "发送成功比例")
        self.card_today = StatsCard("📈", "今日发送", "0", "今天已发送数量")
        self.card_contacts = StatsCard("👥", "联系人", "0", "已导入联系人数")
        
        stats_grid.addWidget(self.card_total, 0, 0)
        stats_grid.addWidget(self.card_success, 0, 1)
        stats_grid.addWidget(self.card_today, 1, 0)
        stats_grid.addWidget(self.card_contacts, 1, 1)
        
        layout.addLayout(stats_grid)
        
        # 运行状态
        status_frame = QFrame()
        status_frame.setObjectName("statusCard")
        status_layout = QVBoxLayout(status_frame)
        status_layout.setContentsMargins(20, 20, 20, 20)
        status_layout.setSpacing(12)
        
        status_header = QHBoxLayout()
        status_title = QLabel("📊 运行状态")
        status_title.setObjectName("sectionTitle")
        status_header.addWidget(status_title)
        
        # 新增状态胶囊（替代进度条文字）
        self.status_badge = QLabel("🟢 准备就绪")
        self.status_badge.setObjectName("statusBadge")
        self.status_badge.setStyleSheet("color: #10B981; background-color: rgba(16, 185, 129, 0.1);")
        status_header.addWidget(self.status_badge)
        
        status_header.addStretch()
        status_layout.addLayout(status_header)
        
        self.dashboard_progress_bar = QProgressBar()
        self.dashboard_progress_bar.setValue(0)
        self.dashboard_progress_bar.setTextVisible(False) # 不显示文字
        self.dashboard_progress_bar.setVisible(False) # 平时隐藏
        status_layout.addWidget(self.dashboard_progress_bar)
        
        # 使用 CodeLogWidget 保持样式一致
        self.log_viewer = CodeLogWidget()
        self.log_viewer.setPlaceholderText("运行日志将显示在这里...")
        self.log_viewer.setMaximumHeight(120)
        status_layout.addWidget(self.log_viewer)
        
        layout.addWidget(status_frame)
        layout.addStretch()
        
        return page

    def create_collect_page(self):
        """创建采集联系人页面 - 增强版"""
        page = QWidget()
        page.setObjectName("contentArea")
        
        scroll = QScrollArea()
        scroll.setObjectName("collectScrollArea")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea#collectScrollArea { background: transparent; border: none; }")
        
        scroll_content = QWidget()
        scroll_content.setObjectName("collectionScrollContent")
        scroll_content.setStyleSheet("QWidget#collectionScrollContent { background: transparent; }")
        layout = QVBoxLayout(scroll_content)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(20)
        
        # 页面标题
        title = QLabel("采集联系人")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        
        # ===== 采集设置卡片 =====
        card = QFrame()
        card.setObjectName("contentCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(24, 24, 24, 24)
        card_layout.setSpacing(16)
        
        # 第1行：采集来源 + 采集数量
        row1 = QHBoxLayout()
        row1.setSpacing(24)
        
        # 采集来源
        source_group = QHBoxLayout()
        source_group.setSpacing(12)
        source_label = QLabel("采集来源")
        source_label.setObjectName("fieldLabel")
        source_label.setMinimumWidth(70)
        
        self.folder_combo = QComboBox()
        self.folder_combo.addItem("📥 收件箱", "inbox")
        self.folder_combo.addItem("📤 已发送", "sent")
        self.folder_combo.addItem("📁 全部文件夹", "all")
        self.folder_combo.setMinimumWidth(140)
        
        source_group.addWidget(source_label)
        source_group.addWidget(self.folder_combo)
        row1.addLayout(source_group)
        
        # 采集数量
        limit_group = QHBoxLayout()
        limit_group.setSpacing(12)
        limit_label = QLabel("采集数量")
        limit_label.setObjectName("fieldLabel")
        limit_label.setMinimumWidth(50)
        
        self.limit_spin = PremiumSpinBox()
        self.limit_spin.setRange(1, 9999999) # Effectively unlimited
        self.limit_spin.setValue(200)
        self.limit_spin.setSuffix(" 封")
        self.limit_spin.setFixedWidth(150) # Precise control
        
        limit_group.addWidget(limit_label)
        limit_group.addWidget(self.limit_spin)
        row1.addLayout(limit_group)
        row1.addStretch()
        card_layout.addLayout(row1)
        
        # 第2行：时间范围
        row2 = QHBoxLayout()
        row2.setSpacing(24)
        
        date_group = QHBoxLayout()
        date_group.setSpacing(12)
        date_label = QLabel("时间范围")
        date_label.setObjectName("fieldLabel")
        date_label.setMinimumWidth(70)
        
        self.date_combo = QComboBox()
        self.date_combo.addItem(" 全部时间", "all")
        self.date_combo.addItem(" 最近 7 天", "7")
        self.date_combo.addItem(" 最近 30 天", "30")
        self.date_combo.addItem(" 最近 90 天", "90")
        self.date_combo.addItem(" 最近 180 天", "180")
        self.date_combo.addItem(" 最近 365 天", "365")
        self.date_combo.setMinimumWidth(140)
        
        date_group.addWidget(date_label)
        date_group.addWidget(self.date_combo)
        row2.addLayout(date_group)
        row2.addStretch()
        card_layout.addLayout(row2)
        
        # 第3行：域名过滤
        row3 = QHBoxLayout()
        row3.setSpacing(12)
        
        filter_label = QLabel("域名过滤")
        filter_label.setObjectName("fieldLabel")
        filter_label.setMinimumWidth(70)
        
        self.include_domain_input = QLineEdit()
        self.include_domain_input.setPlaceholderText("只包含（如: qq.com, 163.com）留空表示全部")
        
        row3.addWidget(filter_label)
        row3.addWidget(self.include_domain_input, 1)
        card_layout.addLayout(row3)
        
        # 第4行：排除域名
        row4 = QHBoxLayout()
        row4.setSpacing(12)
        
        exclude_label = QLabel("排除域名")
        exclude_label.setObjectName("fieldLabel")
        exclude_label.setMinimumWidth(70)
        
        self.exclude_domain_input = QLineEdit()
        self.exclude_domain_input.setPlaceholderText("排除系统邮件（如: noreply.com, system.com）")
        
        row4.addWidget(exclude_label)
        row4.addWidget(self.exclude_domain_input, 1)
        card_layout.addLayout(row4)
        
        # 按钮组
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        
        self.btn_fetch = QPushButton("🚀 开始采集")
        self.btn_fetch.setObjectName("primaryButton")
        self.btn_fetch.clicked.connect(self.start_fetch)
        
        self.btn_stop_fetch = QPushButton("⏹ 停止")
        self.btn_stop_fetch.setObjectName("dangerButton")
        self.btn_stop_fetch.clicked.connect(self.stop_fetch)
        self.btn_stop_fetch.setEnabled(False)
        
        self.btn_save_to_db = QPushButton("💾 保存到联系人库")
        self.btn_save_to_db.setObjectName("primaryButton")
        self.btn_save_to_db.clicked.connect(self.save_to_contact_db)
        self.btn_save_to_db.setEnabled(False)
        
        self.btn_save_contacts = QPushButton("📤 导出文件")
        self.btn_save_contacts.setObjectName("secondaryButton")
        self.btn_save_contacts.clicked.connect(self.save_fetched_contacts)
        self.btn_save_contacts.setEnabled(False)
        
        btn_row.addWidget(self.btn_fetch)
        btn_row.addWidget(self.btn_stop_fetch)
        btn_row.addWidget(self.btn_save_to_db)
        btn_row.addWidget(self.btn_save_contacts)
        btn_row.addStretch()
        card_layout.addLayout(btn_row)
        
        layout.addWidget(card)
        
        # ===== 实时日志卡片（科技感）=====
        log_card = QFrame()
        log_card.setObjectName("contentCard")
        log_layout = QVBoxLayout(log_card)
        log_layout.setContentsMargins(24, 24, 24, 24)
        log_layout.setSpacing(12)
        
        log_header = QHBoxLayout()
        log_title = QLabel("💻 实时日志")
        log_title.setObjectName("sectionTitle")
        log_header.addWidget(log_title)
        
        self.log_status = QLabel("⚪ 待命")
        self.log_status.setObjectName("tipLabel")
        log_header.addWidget(self.log_status)
        log_header.addStretch()
        log_layout.addLayout(log_header)
        
        self.code_log = CodeLogWidget()
        log_layout.addWidget(self.code_log)
        
        layout.addWidget(log_card)
        
        # ===== 结果预览卡片 (Priority High - Moved Up) =====
        preview_card = QFrame()
        preview_card.setObjectName("contentCard")
        preview_layout = QVBoxLayout(preview_card)
        preview_layout.setContentsMargins(24, 24, 24, 24)
        preview_layout.setSpacing(12)
        
        preview_header = QHBoxLayout()
        preview_title = QLabel("📋 采集结果")
        preview_title.setObjectName("sectionTitle")
        preview_header.addWidget(preview_title)
        
        self.result_count_label = QLabel("")
        self.result_count_label.setObjectName("tipLabel")
        preview_header.addWidget(self.result_count_label)
        preview_header.addStretch()
        preview_layout.addLayout(preview_header)
        
        self.fetch_preview = QTextEdit()
        self.fetch_preview.setPlaceholderText("采集到的联系人将显示在这里...\n\n格式：邮箱地址 | 姓名 | 互动次数 | 最后联系时间")
        self.fetch_preview.setReadOnly(True)
        self.fetch_preview.setMinimumHeight(200) # Make it large as requested
        preview_layout.addWidget(self.fetch_preview)
        
        layout.addWidget(preview_card, 2) # Higher stretch to make it big
        
        # ===== 统计图表卡片 (Moved Down) =====
        stats_card = QFrame()
        stats_card.setObjectName("contentCard")
        stats_layout = QVBoxLayout(stats_card)
        stats_layout.setContentsMargins(24, 24, 24, 24)
        stats_layout.setSpacing(16)
        
        stats_title = QLabel("📊 采集统计")
        stats_title.setObjectName("sectionTitle")
        stats_layout.addWidget(stats_title)
        
        # 统计数字行
        stats_row = QHBoxLayout()
        stats_row.setSpacing(32)
        
        # 联系人总数
        self.stat_total = self._create_stat_item("联系人总数", "0")
        stats_row.addLayout(self.stat_total)
        
        # 互动总次数
        self.stat_interactions = self._create_stat_item("互动总次数", "0")
        stats_row.addLayout(self.stat_interactions)
        
        # 平均互动
        self.stat_avg = self._create_stat_item("平均互动", "0")
        stats_row.addLayout(self.stat_avg)
        
        stats_row.addStretch()
        stats_layout.addLayout(stats_row)
        
        # 域名分布图表 (Changed to PieChart)
        chart_label = QLabel("📧 域名分布")
        chart_label.setObjectName("fieldLabel")
        stats_layout.addWidget(chart_label)
        
        self.domain_chart = PieChart() # PIE CHART
        stats_layout.addWidget(self.domain_chart)
        
        # 隐藏的文本标签（用于无数据时）
        self.domain_stats_label = QLabel("暂无数据 - 开始采集后将显示图表")
        self.domain_stats_label.setObjectName("tipLabel")
        self.domain_stats_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.domain_stats_label)
        
        layout.addWidget(stats_card, 1) # Standard stretch
        
        scroll.setWidget(scroll_content)
        
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.addWidget(scroll)
        
        return page
    
    def _create_stat_item(self, label: str, value: str):
        """创建统计项"""
        layout = QVBoxLayout()
        layout.setSpacing(4)
        
        value_label = QLabel(value)
        value_label.setObjectName("statValue")
        value_label.setStyleSheet("font-size: 24px; font-weight: 700;")
        
        text_label = QLabel(label)
        text_label.setObjectName("tipLabel")
        
        layout.addWidget(value_label)
        layout.addWidget(text_label)
        
        # 保存引用以便更新
        layout.value_label = value_label
        
        return layout

    def create_send_page(self):
        """创建批量发送页面"""
        page = QWidget()
        page.setObjectName("contentArea")
        
        scroll = QScrollArea()
        scroll.setObjectName("sendScrollArea")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea#sendScrollArea { background: transparent; border: none; }")
        
        scroll_content = QWidget()
        scroll_content.setObjectName("sendScrollContent")
        scroll_content.setStyleSheet("QWidget#sendScrollContent { background: transparent; }")
        layout = QVBoxLayout(scroll_content)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)
        
        # 页面标题
        title = QLabel("批量发送")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        
        # 邮件设置卡片
        card = QFrame()
        card.setObjectName("contentCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(24, 24, 24, 24)
        card_layout.setSpacing(16)
        
        # 收件人来源选择
        source_label = QLabel("收件人来源")
        source_label.setObjectName("fieldLabel")
        card_layout.addWidget(source_label)
        
        source_row = QHBoxLayout()
        source_row.setSpacing(12)
        
        self.btn_from_contacts = QPushButton("📇 从联系人库选择")
        self.btn_from_contacts.setObjectName("primaryButton")
        self.btn_from_contacts.clicked.connect(self.select_from_contacts)
        source_row.addWidget(self.btn_from_contacts)
        
        self.btn_from_file = QPushButton("📄 导入文件")
        self.btn_from_file.setObjectName("secondaryButton")
        self.btn_from_file.clicked.connect(self.load_contacts_file)
        source_row.addWidget(self.btn_from_file)
        
        source_row.addStretch()
        card_layout.addLayout(source_row)
        
        # 已选联系人显示框
        self.contact_info_frame = QFrame()
        self.contact_info_frame.setStyleSheet("""
            QFrame {
                background: rgba(79, 70, 229, 0.1);
                border: 1px solid rgba(79, 70, 229, 0.3);
                border-radius: 8px;
            }
        """)
        contact_info_layout = QVBoxLayout(self.contact_info_frame)
        contact_info_layout.setContentsMargins(16, 12, 16, 12)
        contact_info_layout.setSpacing(8)
        
        self.contact_count_label = QLabel("📬 尚未选择收件人")
        self.contact_count_label.setStyleSheet("color: #9CA3AF; font-size: 13px;")
        contact_info_layout.addWidget(self.contact_count_label)
        
        self.contact_preview = QTextEdit()
        self.contact_preview.setPlaceholderText("选择联系人后将在此显示...")
        self.contact_preview.setReadOnly(True)
        self.contact_preview.setMaximumHeight(80)
        self.contact_preview.setStyleSheet("""
            QTextEdit {
                background: transparent;
                border: none;
                color: #D1D5DB;
                font-size: 12px;
            }
        """)
        contact_info_layout.addWidget(self.contact_preview)
        
        clear_row = QHBoxLayout()
        clear_row.addStretch()
        self.btn_clear_contacts = QPushButton("清空")
        self.btn_clear_contacts.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: none;
                color: #6B7280;
                padding: 4px 8px;
            }
            QPushButton:hover {
                color: #EF4444;
            }
        """)
        self.btn_clear_contacts.clicked.connect(self.clear_selected_contacts)
        clear_row.addWidget(self.btn_clear_contacts)
        contact_info_layout.addLayout(clear_row)
        
        card_layout.addWidget(self.contact_info_frame)
        
        # 隐藏的联系人路径输入（保持兼容）
        self.contact_path_input = QLineEdit()
        self.contact_path_input.setVisible(False)
        
        # 邮件主题
        row2 = QHBoxLayout()
        row2.setSpacing(12)
        label2 = QLabel("邮件主题")
        label2.setObjectName("fieldLabel")
        label2.setMinimumWidth(90)
        self.subject_input = QLineEdit()
        self.subject_input.setPlaceholderText("输入邮件主题")
        row2.addWidget(label2)
        row2.addWidget(self.subject_input, 1)
        card_layout.addLayout(row2)
        
        # 模板选择
        row3 = QHBoxLayout()
        row3.setSpacing(12)
        label3 = QLabel("邮件模板")
        label3.setObjectName("fieldLabel")
        label3.setMinimumWidth(90)
        
        self.template_combo = QComboBox()
        self.template_combo.addItem("(纯文本 - 不使用模板)", "") # Default option
        try:
            from ..templates import TemplateEngine
            engine = TemplateEngine()
            templates = engine.list_templates()
            for tpl in templates:
                self.template_combo.addItem(f"{tpl['display_name']}", tpl['name'])
        except Exception:
            pass
        
        self.template_combo.currentIndexChanged.connect(self.on_template_changed)
        
        self.btn_config_vars = QPushButton("⚙️ 配置")
        self.btn_config_vars.setObjectName("secondaryButton")
        self.btn_config_vars.setEnabled(False)
        self.btn_config_vars.clicked.connect(self.config_template_variables)
        
        self.btn_preview = QPushButton("👁 预览")
        self.btn_preview.setObjectName("secondaryButton")
        self.btn_preview.setEnabled(False)
        self.btn_preview.clicked.connect(self.preview_template)
        
        row3.addWidget(label3)
        row3.addWidget(self.template_combo, 1)
        row3.addWidget(self.btn_config_vars)
        row3.addWidget(self.btn_preview)
        card_layout.addLayout(row3)
        
        self.template_vars = {}
        
        # 邮件正文
        row4 = QVBoxLayout()
        row4.setSpacing(8)
        
        body_header = QHBoxLayout()
        label4 = QLabel("邮件正文")
        label4.setObjectName("fieldLabel")
        
        from PyQt6.QtWidgets import QToolButton
        self.btn_expand_editor = QToolButton()
        self.btn_expand_editor.setText("📝 放大编辑")
        self.btn_expand_editor.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_expand_editor.setStyleSheet("""
            QToolButton {
                border: none;
                background: transparent;
                color: #818CF8;
                font-weight: bold;
                padding: 4px;
            }
            QToolButton:hover {
                color: #A5B4FC;
                background: rgba(129, 140, 248, 0.1);
                border-radius: 4px;
            }
        """)
        self.btn_expand_editor.clicked.connect(self.open_body_editor)
        
        body_header.addWidget(label4)
        body_header.addStretch()
        body_header.addWidget(self.btn_expand_editor)
        
        self.body_input = QTextEdit()
        self.body_input.setPlaceholderText("输入邮件正文内容...")
        self.body_input.setMaximumHeight(200)
        
        row4.addLayout(body_header)
        row4.addWidget(self.body_input)
        card_layout.addLayout(row4)
        
        # 附件（支持多附件）
        self.attachment_paths = []
        
        attach_header = QHBoxLayout()
        attach_header.setSpacing(12)
        label5 = QLabel("添加附件")
        label5.setObjectName("fieldLabel")
        label5.setMinimumWidth(90)
        
        btn_attach = QPushButton("📎 添加文件")
        btn_attach.setObjectName("secondaryButton")
        btn_attach.clicked.connect(self.select_attachment)
        btn_remove = QPushButton("➖ 移除选中")
        btn_remove.setObjectName("secondaryButton")
        btn_remove.clicked.connect(self.remove_selected_attachment)
        btn_clear_attach = QPushButton("🗑 全部清除")
        btn_clear_attach.setObjectName("secondaryButton")
        btn_clear_attach.clicked.connect(self.clear_all_attachments)
        
        attach_header.addWidget(label5)
        attach_header.addStretch()
        attach_header.addWidget(btn_attach)
        attach_header.addWidget(btn_remove)
        attach_header.addWidget(btn_clear_attach)
        card_layout.addLayout(attach_header)
        
        self.attach_list_widget = QListWidget()
        self.attach_list_widget.setMaximumHeight(90)
        self.attach_list_widget.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        self.attach_list_widget.setStyleSheet("""
            QListWidget {
                background-color: #1F2133;
                border: 1px solid #3D3D55;
                border-radius: 8px;
                color: #D1D5DB;
                font-size: 12px;
                padding: 4px;
            }
            QListWidget::item {
                padding: 3px 6px;
            }
            QListWidget::item:selected {
                background-color: rgba(79, 70, 229, 0.3);
            }
        """)
        self.attach_list_widget.setVisible(False)  # 无附件时隐藏
        card_layout.addWidget(self.attach_list_widget)
        
        # 隐藏的兼容字段
        self.attach_path_input = QLineEdit()
        self.attach_path_input.setVisible(False)
        
        # 发送设置
        row6 = QHBoxLayout()
        row6.setSpacing(16)
        
        label6 = QLabel("批次设置")
        label6.setObjectName("fieldLabel")
        label6.setMinimumWidth(90)
        
        self.batch_size_spin = PremiumSpinBox()
        self.batch_size_spin.setRange(1, 100)
        self.batch_size_spin.setValue(10)
        self.batch_size_spin.setPrefix("每批 ")
        self.batch_size_spin.setSuffix(" 封") 
        self.batch_size_spin.setFixedWidth(150)
        
        self.batch_interval_spin = PremiumSpinBox()
        self.batch_interval_spin.setRange(10, 300)
        self.batch_interval_spin.setValue(20)
        self.batch_interval_spin.setPrefix("间隔 ")
        self.batch_interval_spin.setSuffix(" 秒")
        self.batch_interval_spin.setFixedWidth(150)
        
        row6.addWidget(label6)
        row6.addWidget(self.batch_size_spin)
        row6.addWidget(self.batch_interval_spin)
        row6.addStretch()
        card_layout.addLayout(row6)
        
        layout.addWidget(card)
        
        # 发送按钮
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        
        self.btn_send = QPushButton("📤 开始发送")
        self.btn_send.setObjectName("primaryButton")
        self.btn_send.setMinimumHeight(44)
        self.btn_send.clicked.connect(self.start_send)
        
        self.btn_stop_send = QPushButton("⏹ 停止发送")
        self.btn_stop_send.setObjectName("dangerButton")
        self.btn_stop_send.setMinimumHeight(44)
        self.btn_stop_send.clicked.connect(self.stop_send)
        self.btn_stop_send.clicked.connect(self.stop_send)
        # self.btn_stop_send.setEnabled(False)  # User wants to click it for fun
        
        btn_row.addWidget(self.btn_send)
        btn_row.addWidget(self.btn_stop_send)
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("准备发送...")
        self.progress_bar.setMinimumHeight(28)
        layout.addWidget(self.progress_bar)
        
        # ===== 实时日志卡片 =====
        log_card = QFrame()
        log_card.setObjectName("contentCard")
        log_layout = QVBoxLayout(log_card)
        log_layout.setContentsMargins(24, 24, 24, 24)
        log_layout.setSpacing(12)
        
        log_header = QHBoxLayout()
        log_title = QLabel("💬 发送日志")
        log_title.setObjectName("sectionTitle")
        log_header.addWidget(log_title)
        
        self.send_status_label = QLabel("⚪ 待发送")
        self.send_status_label.setObjectName("tipLabel")
        log_header.addWidget(self.send_status_label)
        log_header.addStretch()
        log_layout.addLayout(log_header)
        
        self.send_log = CodeLogWidget()
        log_layout.addWidget(self.send_log)
        
        layout.addWidget(log_card)
        
        layout.addStretch()
        
        scroll.setWidget(scroll_content)
        
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.addWidget(scroll)
        
        return page

    def create_settings_page(self):
        """创建设置页面"""
        page = QWidget()
        page.setObjectName("contentArea")
        layout = QVBoxLayout(page)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(24)
        
        # 页面标题
        title = QLabel("账号设置")
        title.setObjectName("pageTitle")
        layout.addWidget(title)
        
        # 账号配置卡片
        card = QFrame()
        card.setObjectName("contentCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(24, 24, 24, 24)
        card_layout.setSpacing(20)
        
        # 邮箱账号
        row1 = QHBoxLayout()
        row1.setSpacing(12)
        label1 = QLabel("📧 邮箱账号")
        label1.setObjectName("fieldLabel")
        label1.setMinimumWidth(100)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("your_email@163.com")
        self.email_input.textChanged.connect(self.on_email_input_changed)
        row1.addWidget(label1)
        row1.addWidget(self.email_input, 1)
        card_layout.addLayout(row1)
        
        # 授权码
        row2 = QHBoxLayout()
        row2.setSpacing(12)
        label2 = QLabel("🔑 授权码")
        label2.setObjectName("fieldLabel")
        label2.setMinimumWidth(100)
        self.pwd_input = QLineEdit()
        self.pwd_input.setPlaceholderText("授权码（非登录密码）")
        self.pwd_input.setEchoMode(QLineEdit.EchoMode.Password)
        row2.addWidget(label2)
        row2.addWidget(self.pwd_input, 1)
        card_layout.addLayout(row2)
        
        # 邮箱类型选择
        row3 = QHBoxLayout()
        row3.setSpacing(12)
        label3 = QLabel("📮 邮箱类型")
        label3.setObjectName("fieldLabel")
        label3.setMinimumWidth(100)
        
        self.email_type_combo = QComboBox()
        self.email_type_combo.addItem("🔄 自动识别", "auto")
        self.email_type_combo.addItem(" 163 邮箱", "163.com")
        self.email_type_combo.addItem(" 126 邮箱", "126.com")
        self.email_type_combo.addItem(" QQ 邮箱", "qq.com")
        self.email_type_combo.addItem(" 阿里云邮箱", "aliyun.com")
        self.email_type_combo.addItem(" 新浪邮箱", "sina.com")
        self.email_type_combo.addItem(" Gmail", "gmail.com")
        self.email_type_combo.addItem(" Outlook", "outlook.com")
        self.email_type_combo.currentIndexChanged.connect(self.on_email_type_changed)
        
        self.server_info_label = QLabel("")
        self.server_info_label.setObjectName("tipLabel")
        
        row3.addWidget(label3)
        row3.addWidget(self.email_type_combo)
        row3.addWidget(self.server_info_label, 1)
        card_layout.addLayout(row3)
        
        # 更新服务器信息显示
        self.on_email_type_changed()
        
        # 提示
        tip = QLabel("💡 提示：授权码需要在邮箱设置中开启SMTP服务后获取")
        tip.setObjectName("tipLabel")
        card_layout.addWidget(tip)
        
        # 保存按钮
        btn_save = QPushButton("💾 保存设置")
        btn_save.setObjectName("primaryButton")
        btn_save.clicked.connect(self.save_config)
        card_layout.addWidget(btn_save, alignment=Qt.AlignmentFlag.AlignLeft)
        
        layout.addWidget(card)
        layout.addStretch()
        
        return page

    def switch_page(self, index: int):
        """切换页面"""
        self.pages.setCurrentIndex(index)
        for i, btn in enumerate(self.nav_buttons):
            btn.setChecked(i == index)

    # ========== 功能方法 ==========
    def log(self, message, target='collect'):
        """
        写入日志
        target: 'collect' - 采集页面日志, 'send' - 发送页面日志, 'both' - 两个都写
        """
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        formatted_msg = f"[{timestamp}] {message}"
        
        if target in ('collect', 'both'):
            self.log_viewer.append(formatted_msg)
            cursor = self.log_viewer.textCursor()
            cursor.movePosition(cursor.MoveOperation.End)
            self.log_viewer.setTextCursor(cursor)
        
        if target in ('send', 'both') and hasattr(self, 'send_log'):
            self.send_log.append(formatted_msg)

    def on_email_type_changed(self):
        """邮箱类型选择变更"""
        from ..core.email_config import get_email_config, EMAIL_SERVERS
        
        email_type = self.email_type_combo.currentData()
        
        if email_type == "auto":
            # 自动识别模式，根据输入的邮箱地址显示
            email = self.email_input.text().strip()
            if email and '@' in email:
                config = get_email_config(email)
                self.server_info_label.setText(f"SMTP: {config['smtp']}")
            else:
                self.server_info_label.setText("输入邮箱后自动识别服务器")
        else:
            # 手动选择模式
            if email_type in EMAIL_SERVERS:
                config = EMAIL_SERVERS[email_type]
                self.server_info_label.setText(f"SMTP: {config['smtp']}")
            else:
                self.server_info_label.setText("")
    
    def on_email_input_changed(self, text: str):
        """邮箱输入变化时更新服务器信息"""
        if self.email_type_combo.currentData() == "auto":
            self.on_email_type_changed()
    
    def get_selected_email_type(self) -> str:
        """获取选择的邮箱类型"""
        return self.email_type_combo.currentData()

    def load_config(self):
        self.email_input.setText(self.config_manager.settings.value("email", ""))
        self.pwd_input.setText(self.config_manager.settings.value("pwd", ""))
        self.contact_path_input.setText(self.config_manager.settings.value("last_contact_file", ""))
        self.subject_input.setText(self.config_manager.settings.value("last_subject", ""))
        self.body_input.setPlainText(self.config_manager.settings.value("last_body", ""))
        # 加载多附件
        import json as _json
        saved_attachments = self.config_manager.settings.value("last_attachments", "[]")
        try:
            self.attachment_paths = _json.loads(saved_attachments) if saved_attachments else []
        except Exception:
            self.attachment_paths = []
        self._refresh_attach_list()
        
        # 加载邮箱类型
        saved_type = self.config_manager.settings.value("email_type", "auto")
        index = self.email_type_combo.findData(saved_type)
        if index >= 0:
            self.email_type_combo.setCurrentIndex(index)
        self.on_email_type_changed()

    def save_config(self):
        """保存配置（静默保存，不显示弹窗）"""
        self.config_manager.settings.setValue("email", self.email_input.text())
        self.config_manager.settings.setValue("pwd", self.pwd_input.text())
        self.config_manager.settings.setValue("email_type", self.email_type_combo.currentData())
        self.config_manager.settings.setValue("last_contact_file", self.contact_path_input.text())
        self.config_manager.settings.setValue("last_subject", self.subject_input.text())
        self.config_manager.settings.setValue("last_body", self.body_input.toPlainText())
        import json as _json
        self.config_manager.settings.setValue("last_attachments", _json.dumps(self.attachment_paths))

    def load_stats(self):
        self.total_sent = self.config_manager.settings.value('stats/total_sent', 0, type=int)
        self.success_count = self.config_manager.settings.value('stats/success_count', 0, type=int)
        today = datetime.now().strftime('%Y-%m-%d')
        saved_date = self.config_manager.settings.value('stats/today_date', '')
        if saved_date == today:
            self.today_sent = self.config_manager.settings.value('stats/today_sent', 0, type=int)
        else:
            self.today_sent = 0
        self.update_dashboard()

    def save_stats(self):
        self.config_manager.settings.setValue('stats/total_sent', self.total_sent)
        self.config_manager.settings.setValue('stats/success_count', self.success_count)
        self.config_manager.settings.setValue('stats/today_sent', self.today_sent)
        self.config_manager.settings.setValue('stats/today_date', datetime.now().strftime('%Y-%m-%d'))

    def update_dashboard(self):
        self.card_total.update_value(str(self.total_sent))
        rate = (self.success_count / self.total_sent * 100) if self.total_sent > 0 else 0
        self.card_success.update_value(f"{rate:.0f}%")
        self.card_today.update_value(str(self.today_sent))
        self.card_contacts.update_value(str(len(self.contacts_data)))

    # ========== 采集功能（增强版）==========
    def start_fetch(self):
        user = self.email_input.text().strip()
        pwd = self.pwd_input.text().strip()
        
        if not user or not pwd:
            QMessageBox.warning(self, "⚠️ 缺少信息", "请先在「账号设置」中填写邮箱账号和授权码")
            self.switch_page(3)
            return

        self.fetch_preview.clear()
        self.log("开始采集联系人...")
        self.btn_fetch.setEnabled(False)
        self.btn_stop_fetch.setEnabled(True)
        self.btn_save_contacts.setEnabled(False)
        self.contacts_data = []
        
        # 重置统计显示
        self.stat_total.value_label.setText("0")
        self.stat_interactions.value_label.setText("0")
        self.stat_avg.value_label.setText("0")
        self.domain_stats_label.setText("等待数据...")
        self.domain_chart.set_data({})
        self.result_count_label.setText("")
        
        # 初始化代码日志
        self.code_log.clear_logs()
        self.log_status.setText("🟢 运行中")
        self.log_status.setStyleSheet("color: #10B981;")
        
        # 科技感日志
        self.code_log.add_log(">>> INIT CONTACT_FETCHER MODULE", "info")
        self.code_log.add_log(f"USER: {user}", "data")
        
        folder = self.folder_combo.currentData()
        date_range = self.date_combo.currentData()
        self.code_log.add_log(f"CONFIG: folder={folder}, date_range={date_range}", "data")
        
        # 解析过滤条件
        include_domains = []
        exclude_domains = []
        
        include_text = self.include_domain_input.text().strip()
        if include_text:
            include_domains = [d.strip() for d in include_text.split(',') if d.strip()]
            self.code_log.add_log(f"FILTER_INCLUDE: {include_domains}", "data")
        
        exclude_text = self.exclude_domain_input.text().strip()
        if exclude_text:
            exclude_domains = [d.strip() for d in exclude_text.split(',') if d.strip()]
            self.code_log.add_log(f"FILTER_EXCLUDE: {exclude_domains}", "data")
        
        self.code_log.add_log("CONNECTING TO IMAP SERVER...", "progress")
        
        # 构建采集选项
        options = {
            'folder': folder,
            'date_range': date_range,
            'include_domains': include_domains,
            'exclude_domains': exclude_domains,
        }

        self.status_badge.setText("🔵 正在采集...")
        self.status_badge.setStyleSheet("color: #6B7FEB; background-color: rgba(107, 127, 235, 0.1);")
        self.dashboard_progress_bar.setVisible(True)
        self.dashboard_progress_bar.setRange(0, 0) # 采集是不定进度的

        self.fetch_thread = ContactFetcher(user, pwd, self.limit_spin.value(), options=options)
        self.fetch_thread.progress.connect(self.on_fetch_progress)
        self.fetch_thread.result.connect(self.on_fetch_result)
        self.fetch_thread.error.connect(self.on_fetch_error)
        self.fetch_thread.finished.connect(self.on_fetch_finished)
        self.fetch_thread.stats.connect(self.on_fetch_stats)
        self.fetch_thread.contacts_ready.connect(self.on_contacts_ready)
        self.fetch_thread.start()

    def stop_fetch(self):
        if self.fetch_thread and self.fetch_thread.isRunning():
            self.fetch_thread.stop()
            self.log("正在停止采集...")
            self.code_log.add_log(">>> USER_INTERRUPT: STOPPING...", "warning")
            self.log_status.setText("🟡 停止中")
            self.log_status.setStyleSheet("color: #F59E0B;")
            
            # Dashboard status
            self.status_badge.setText("🟡 停止中")
            self.status_badge.setStyleSheet("color: #F59E0B; background-color: rgba(245, 158, 11, 0.1);")

    def on_fetch_progress(self, count: int, email: str):
        self.result_count_label.setText(f"已发现 {count} 个联系人")
        # 每10个联系人记录一次日志
        if count % 10 == 0 or count <= 5:
            self.code_log.add_log(f"FOUND [{count}]: {email}", "success")

    def on_fetch_result(self, message: str):
        self.log(message)
        # 解析消息类型来添加科技感日志
        if "连接" in message or "登录" in message:
            self.code_log.add_log(f"IMAP_AUTH: {message}", "success")
        elif "扫描" in message or "文件夹" in message:
            self.code_log.add_log(f"SCANNING: {message}", "progress")
        elif "处理" in message:
            self.code_log.add_log(f"PROCESS: {message}", "info")
        else:
            self.code_log.add_log(message, "info")

    def on_fetch_error(self, error: str):
        self.log(f"❌ 错误: {error}")
        self.code_log.add_log(f"ERROR: {error}", "error")
        self.log_status.setText("🔴 错误")
        self.log_status.setStyleSheet("color: #EF4444;")
        
        # Dashboard status error
        self.status_badge.setText("🔴 发生错误")
        self.status_badge.setStyleSheet("color: #EF4444; background-color: rgba(239, 68, 68, 0.1);")
        self.dashboard_progress_bar.setVisible(False)
        QMessageBox.critical(self, "❌ 采集失败", error)
    
    def on_fetch_stats(self, stats: dict):
        """更新统计数据显示"""
        total = stats.get('total_contacts', 0)
        interactions = stats.get('total_interactions', 0)
        avg = stats.get('avg_interactions', 0)
        
        self.stat_total.value_label.setText(str(total))
        self.stat_interactions.value_label.setText(str(interactions))
        self.stat_avg.value_label.setText(str(avg))
        
        # 域名分布图表
        domain_dist = stats.get('domain_distribution', {})
        if domain_dist:
            self.domain_chart.set_data(domain_dist)
            self.domain_stats_label.setText("")  # 隐藏文本
            self.code_log.add_log(f"STATS: {total} contacts, {interactions} interactions", "data")
        else:
            self.domain_stats_label.setText("暂无数据")
    
    def on_contacts_ready(self, contacts: list):
        """联系人列表准备完毕"""
        self.contacts_data = contacts
        self.fetch_preview.clear()
        
        self.code_log.add_log(f">>> CONTACTS_READY: {len(contacts)} records", "success")
        
        # 显示联系人列表（带详细信息）
        for c in contacts:
            line = f"{c['email']}"
            if c.get('name'):
                line += f"  |  {c['name']}"
            line += f"  |  互动 {c.get('count', 1)} 次"
            if c.get('last_contact'):
                line += f"  |  最后: {c['last_contact']}"
            self.fetch_preview.append(line)
        
        self.result_count_label.setText(f"共 {len(contacts)} 个联系人")
        self.update_dashboard()

    def on_fetch_finished(self):
        self.btn_fetch.setEnabled(True)
        self.btn_stop_fetch.setEnabled(False)
        if self.contacts_data:
            self.btn_save_to_db.setEnabled(True)
            self.btn_save_contacts.setEnabled(True)
        self.log(f"✅ 采集完成，共 {len(self.contacts_data)} 个联系人")
        
        # 完成日志
        self.code_log.add_log(">>> FETCH_COMPLETE", "success")
        self.code_log.add_log(f"TOTAL_CONTACTS: {len(self.contacts_data)}", "data")
        self.code_log.add_log("SESSION_CLOSED", "info")
        
        self.log_status.setText("⚪ 完成")
        self.log_status.setStyleSheet("color: #9CA3AF;")
        
        # Dashboard status reset
        self.status_badge.setText("🟢 准备就绪")
        self.status_badge.setStyleSheet("color: #10B981; background-color: rgba(16, 185, 129, 0.1);")
        self.dashboard_progress_bar.setVisible(False)
        
        self.update_dashboard()
    
    def save_to_contact_db(self):
        """保存采集到的联系人到数据库"""
        if not self.contacts_data:
            return
        
        # 让用户选择分组
        from ..core import ContactDatabase
        db = ContactDatabase()
        groups = db.get_groups()
        
        # 创建分组选择对话框
        from PyQt6.QtWidgets import QInputDialog
        group_names = ["未分组"] + [f"{g['icon']} {g['name']}" for g in groups]
        
        choice, ok = QInputDialog.getItem(
            self, "选择分组", 
            f"将 {len(self.contacts_data)} 个联系人保存到哪个分组？",
            group_names, 0, False
        )
        
        if not ok:
            return
        
        # 确定分组ID
        group_id = None
        if choice != "未分组":
            for g in groups:
                if f"{g['icon']} {g['name']}" == choice:
                    group_id = g['id']
                    break
        
        # 保存到数据库
        count = self.contact_manager.save_fetched_contacts(self.contacts_data, group_id)
        
        self.log(f"✅ 已保存 {count} 个联系人到联系人库")
        QMessageBox.information(
            self, "✅ 保存成功", 
            f"已保存 {count} 个联系人到联系人库\n\n点击「联系人管理」查看"
        )
        
        # 刷新联系人管理页面
        self.contact_manager.load_data()
    
    def on_contacts_selected_for_send(self, contacts: list):
        """从联系人管理页面选择联系人发送"""
        self.contacts_data = contacts
        self.update_dashboard()
        
        # 切换到发送页面
        self.switch_page(3)
        
        # 更新联系人显示
        self.update_contact_display(contacts)
        
        QMessageBox.information(
            self, "已加载联系人",
            f"已加载 {len(contacts)} 个联系人，请填写邮件内容后发送"
        )
    
    def on_quick_send_requested(self, contact: dict):
        """快捷发送邮件弹窗"""
        sender_email = self.email_input.text().strip()
        sender_pwd = self.pwd_input.text().strip()
        
        dialog = QuickSendDialog(
            contact=contact,
            sender_email=sender_email,
            sender_pwd=sender_pwd,
            parent=self
        )
        dialog.exec()
        
        # 刷新联系人管理页面（更新发送记录）
        self.contact_manager.load_data()
    
    def select_from_contacts(self):
        """从联系人库选择收件人"""
        dialog = ContactSelectDialog(parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            contacts = dialog.get_selected_contacts()
            if contacts:
                self.contacts_data = contacts
                self.update_contact_display(contacts)
                self.update_dashboard()
    
    def update_contact_display(self, contacts: list):
        """更新联系人显示"""
        count = len(contacts)
        self.contact_count_label.setText(f"📬 已选择 {count} 个收件人")
        
        # 显示预览（最多显示5个）
        preview_lines = []
        for c in contacts[:5]:
            line = c['email']
            if c.get('name'):
                line += f" ({c['name']})"
            preview_lines.append(line)
        
        if count > 5:
            preview_lines.append(f"... 等 {count} 人")
        
        self.contact_preview.setPlainText("\n".join(preview_lines))
        self.contact_path_input.setText(f"[已选择 {count} 个联系人]")
    
    def clear_selected_contacts(self):
        """清空已选联系人"""
        self.contacts_data = []
        self.contact_count_label.setText("📬 尚未选择收件人")
        self.contact_preview.clear()
        self.contact_path_input.clear()
        self.update_dashboard()

    def save_fetched_contacts(self):
        if not self.contacts_data:
            return
        
        path, selected_filter = QFileDialog.getSaveFileName(
            self, 
            "导出联系人", 
            "", 
            "CSV文件 (*.csv);;Excel文件 (*.xlsx);;文本文件 (*.txt);;通讯录文件 (*.vcf)"
        )
        
        if not path:
            return
        
        try:
            if path.endswith('.csv') or 'CSV' in selected_filter:
                self._export_csv(path)
            elif path.endswith('.xlsx') or 'Excel' in selected_filter:
                self._export_excel(path)
            elif path.endswith('.vcf') or '通讯录' in selected_filter:
                self._export_vcf(path)
            else:
                self._export_txt(path)
            
            self.log(f"✅ 联系人已保存到: {path}")
            QMessageBox.information(self, "✅ 导出成功", f"已导出 {len(self.contacts_data)} 个联系人到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "❌ 导出失败", f"导出失败: {str(e)}")
    
    def _export_csv(self, path: str):
        """导出为 CSV"""
        import csv
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['邮箱', '姓名', '互动次数', '最后联系', '首次联系'])
            for c in self.contacts_data:
                writer.writerow([
                    c.get('email', ''),
                    c.get('name', ''),
                    c.get('count', 1),
                    c.get('last_contact', ''),
                    c.get('first_contact', ''),
                ])
    
    def _export_excel(self, path: str):
        """导出为 Excel"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "联系人"
            
            # 表头
            headers = ['邮箱', '姓名', '互动次数', '最后联系', '首次联系']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 数据
            for row, c in enumerate(self.contacts_data, 2):
                ws.cell(row=row, column=1, value=c.get('email', ''))
                ws.cell(row=row, column=2, value=c.get('name', ''))
                ws.cell(row=row, column=3, value=c.get('count', 1))
                ws.cell(row=row, column=4, value=c.get('last_contact', ''))
                ws.cell(row=row, column=5, value=c.get('first_contact', ''))
            
            wb.save(path)
        except ImportError:
            # 没有 openpyxl，降级为 CSV
            csv_path = path.replace('.xlsx', '.csv')
            self._export_csv(csv_path)
            raise Exception(f"未安装 openpyxl，已导出为 CSV: {csv_path}")
    
    def _export_vcf(self, path: str):
        """导出为 VCF 通讯录格式"""
        with open(path, 'w', encoding='utf-8') as f:
            for c in self.contacts_data:
                email = c.get('email', '')
                name = c.get('name', '') or email.split('@')[0]
                
                f.write("BEGIN:VCARD\n")
                f.write("VERSION:3.0\n")
                f.write(f"FN:{name}\n")
                f.write(f"EMAIL:{email}\n")
                f.write("END:VCARD\n")
    
    def _export_txt(self, path: str):
        """导出为纯文本"""
        with open(path, 'w', encoding='utf-8') as f:
            for c in self.contacts_data:
                f.write(c.get('email', '') + '\n')

    # ========== 发送功能 ==========
    def load_contacts_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择联系人文件", "", "支持的文件 (*.txt *.csv);;文本文件 (*.txt);;CSV文件 (*.csv)")
        if path:
            self.contact_path_input.setText(path)
            contacts = read_contacts(path)
            self.contacts_data = contacts
            self.update_dashboard()
            self.log(f"✅ 已加载 {len(contacts)} 个联系人")

    def on_template_changed(self, index):
        """模板选择变更"""
        template_name = self.template_combo.currentData()
        is_template_mode = bool(template_name)
        
        self.btn_config_vars.setEnabled(is_template_mode)
        self.btn_preview.setEnabled(is_template_mode)
        self.body_input.setEnabled(not is_template_mode)
        
        if is_template_mode:
             self.body_input.setPlaceholderText("⚠️ 已选择HTML模板，正文输入框将被忽略（请点击「配置」按钮修改变量）")
        else:
             self.body_input.setPlaceholderText("输入邮件正文内容...")

    def config_template_variables(self):
        from .variable_config_dialog import VariableConfigDialog
        template_name = self.template_combo.currentData()
        if not template_name:
            QMessageBox.warning(self, "⚠️ 提示", "请先选择一个模板")
            return
        if not isinstance(self.template_vars, dict):
             self.template_vars = {}
        dialog = VariableConfigDialog(self.template_vars, [], self)
        if dialog.exec():
            self.template_vars = dialog.get_variables()
            self.log(f"✅ 模板变量已配置: {len(self.template_vars)} 个")

    def open_body_editor(self):
        """打开放大版正文编辑器"""
        from .custom_widgets import EmailEditorDialog
        
        current_text = self.body_input.toPlainText()
        dialog = EmailEditorDialog(current_text, self)
        
        if dialog.exec():
            new_text = dialog.get_text()
            self.body_input.setPlainText(new_text)

    def preview_template(self):
        from .template_preview import TemplatePreviewDialog
        template_name = self.template_combo.currentData()
        if not template_name:
            QMessageBox.warning(self, "⚠️ 提示", "请先选择一个模板")
            return
        try:
            from ..templates import TemplateEngine
            engine = TemplateEngine()
            preview_vars = {**self.template_vars, 'recipient_name': '预览用户', 'recipient_email': 'preview@example.com'}
            html = engine.render(template_name, preview_vars)
            dialog = TemplatePreviewDialog(html, self)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "❌ 预览失败", f"模板预览失败:\n\n{str(e)}")

    def select_attachment(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择附件（可多选）", "", "图片和PDF (*.png *.jpg *.jpeg *.gif *.pdf);;Office文件 (*.doc *.docx *.xls *.xlsx *.ppt *.pptx);;所有文件 (*)")
        if paths:
            for p in paths:
                if p not in self.attachment_paths:
                    self.attachment_paths.append(p)
            self._refresh_attach_list()

    def remove_selected_attachment(self):
        """移除选中的附件"""
        selected = self.attach_list_widget.selectedItems()
        for item in selected:
            row = self.attach_list_widget.row(item)
            if 0 <= row < len(self.attachment_paths):
                self.attachment_paths.pop(row)
        self._refresh_attach_list()

    def clear_all_attachments(self):
        """清除所有附件"""
        self.attachment_paths.clear()
        self._refresh_attach_list()

    def _refresh_attach_list(self):
        """刷新附件列表显示"""
        self.attach_list_widget.clear()
        import os
        for p in self.attachment_paths:
            filename = os.path.basename(p)
            self.attach_list_widget.addItem(f"📎 {filename}")
        self.attach_list_widget.setVisible(len(self.attachment_paths) > 0)

    def start_send(self):
        user = self.email_input.text().strip()
        pwd = self.pwd_input.text().strip()
        subject = self.subject_input.text().strip()

        if not user or not pwd:
            QMessageBox.warning(self, "⚠️ 缺少信息", "请先在「账号设置」中填写邮箱账号和授权码")
            self.switch_page(3)
            return

        if not self.contacts_data:
            QMessageBox.warning(self, "⚠️ 缺少联系人", "请先加载联系人列表")
            return

        if not subject:
            QMessageBox.warning(self, "⚠️ 缺少主题", "请填写邮件主题")
            return

        body = ""
        html_body = None
        
        template_name = self.template_combo.currentData()
        if template_name: # Use template
            try:
                from ..templates import TemplateEngine
                engine = TemplateEngine()
                html_body = engine.render(template_name, self.template_vars)
            except Exception as e:
                QMessageBox.critical(self, "❌ 模板错误", f"渲染模板失败:\n{str(e)}")
                return
        else: # Plain text
            body = self.body_input.toPlainText().strip()
            if not body:
                QMessageBox.warning(self, "⚠️ 缺少正文", "请填写邮件正文")
                return

        attachments = self.attachment_paths if self.attachment_paths else None

        self.save_config()
        
        # 清空并初始化发送日志
        self.send_log.clear()
        self.send_status_label.setText("🟢 发送中...")
        self.log(f"📧 开始发送邮件，共 {len(self.contacts_data)} 位收件人", 'send')
        self.log(f"📦 批次设置: 每批 {self.batch_size_spin.value()} 封，间隔 {self.batch_interval_spin.value()} 秒", 'send')
        if attachments:
            self.log(f"📎 附件: {len(attachments)} 个文件", 'send')
        
        self.btn_send.setEnabled(False)
        self.btn_stop_send.setEnabled(True)
        self.progress_bar.setMaximum(len(self.contacts_data))
        self.progress_bar.setValue(0)
        
        # Dashboard status update
        self.status_badge.setText("🔵 正在发送...")
        self.status_badge.setStyleSheet("color: #6B7FEB; background-color: rgba(107, 127, 235, 0.1);")
        self.dashboard_progress_bar.setVisible(True)
        self.dashboard_progress_bar.setRange(0, len(self.contacts_data))
        self.dashboard_progress_bar.setValue(0)

        self.send_thread = EmailSender(
            user, pwd, self.contacts_data, subject, body, attachments,
            self.batch_size_spin.value(), self.batch_interval_spin.value(),
            html_body
        )
        self.send_thread.progress.connect(self.on_send_progress)
        self.send_thread.result.connect(self.on_send_result)
        self.send_thread.error.connect(self.on_send_error)
        self.send_thread.finished.connect(self.on_send_finished)
        self.send_thread.batch_done.connect(self.on_batch_done)
        self.send_thread.wait_progress.connect(self.on_wait_progress)
        self.send_thread.start()

    def stop_send(self):
        if self.send_thread and self.send_thread.isRunning():
            self.send_thread.stop()
            self.log("⏹ 正在停止发送...", 'send')
            self.send_status_label.setText("🟡 停止中...")

    def on_send_progress(self, current: int, total: int, email: str):
        self.progress_bar.setValue(current)
        self.progress_bar.setFormat(f"发送中 {current}/{total}")
        
        # Sync dashboard progress
        self.dashboard_progress_bar.setValue(current)
        self.log(f"✅ [{current}/{total}] 已发送: {email}", 'send')

    def on_send_result(self, success: int, failed: int):
        self.total_sent += success + failed
        self.success_count += success
        self.today_sent += success + failed
        self.save_stats()
        self.update_dashboard()
        self.log(f"📊 发送统计: 成功 {success} 封, 失败 {failed} 封", 'send')

    def on_send_error(self, error: str):
        self.log(f"❌ {error}", 'send')
        
        # Dashboard status error
        self.status_badge.setText("🔴 发送错误")
        self.status_badge.setStyleSheet("color: #EF4444; background-color: rgba(239, 68, 68, 0.1);")
        self.dashboard_progress_bar.setVisible(False)

    def on_batch_done(self, batch_num: int, wait_time: int):
        """批次完成，等待中"""
        self.log(f"📦 第 {batch_num} 批完成，等待 {wait_time} 秒后继续...", 'send')
        self.progress_bar.setFormat(f"等待中... {wait_time}s")
        
    def on_wait_progress(self, remaining: int):
        """倒计时更新"""
        self.progress_bar.setFormat(f"⏳ 等待中... {remaining}s")
        self.send_status_label.setText(f"⏳ 等待下一批 {remaining}s")

    def on_send_finished(self):
        self.btn_send.setEnabled(True)
        self.btn_stop_send.setEnabled(False)
        self.progress_bar.setFormat("✅ 发送完成")
        self.send_status_label.setText("✅ 已完成")
        
        # Dashboard status reset
        self.status_badge.setText("🟢 发送完成")
        self.status_badge.setStyleSheet("color: #10B981; background-color: rgba(16, 185, 129, 0.1);")
        self.dashboard_progress_bar.setVisible(False)
        self.log("🎉 邮件发送任务完成！", 'send')
        self.update_dashboard()

    # 主题切换
    def apply_theme(self, theme=None):
        """应用统一的紫色主题"""
        self.setStyleSheet(PURPLE_THEME)
        
        # 强制子组件使用暗色模式逻辑
        if hasattr(self, 'contact_manager'):
            # Purple Theme 本质上是 Dark Mode
            self.contact_manager.update_theme(True)

    # methods removed: toggle_theme, update_theme_button_text


